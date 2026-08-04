import os
import re
import glob
from copy import copy
import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
from openpyxl.worksheet.properties import PageSetupProperties
from PIL import Image as PILImage, ImageOps

try:
    import imagehash
    HAS_IMAGEHASH = True
except ImportError:
    HAS_IMAGEHASH = False

IMG_W_PX, IMG_H_PX = 1200, 900
BOX_WIDTH_EMU = 3155040
BOX_HEIGHT_EMU = 3294720
HASH_DUPLICATE_THRESHOLD = 8

# 사진 1행이 차지하는 실제 엑셀 행 번호 패턴 (템플릿 구조 기준)
# [ (공사명행, 공종행, 세부설명행, 사진행), ... ]
# 페이지(제목+사진 2장) 하나가 14행. 템플릿에 원래부터 만들어져 있는 건 6페이지
# (12장) 뿐이지만, 그보다 더 필요하면 clone_page()로 페이지를 복제해서 늘림.
PAGE_ROWS = 14
PAGE_HEADER_FIRST_ROW = 1
KNOWN_GOOD_PAGES = 6  # 템플릿에 원래(또는 이미 복구되어) 정상적으로 만들어진 페이지 수


def generate_block_rows(count):
    """1페이지=14행, 페이지마다 블록 2개(사진 2장) 규칙에 따라 처음 count개의
    (공사명행, 공종행, 세부설명행, 사진행) 튜플을 계산해서 반환.
    페이지 시작행 p 기준: 사진1행=p+2, 공사명1=p+4, 공종1=p+5, 세부설명1=p+6,
                         사진2행=p+9, 공사명2=p+11, 공종2=p+12, 세부설명2=p+13"""
    blocks = []
    page = PAGE_HEADER_FIRST_ROW
    while len(blocks) < count:
        blocks.append((page + 4, page + 5, page + 6, page + 2))
        blocks.append((page + 11, page + 12, page + 13, page + 9))
        page += PAGE_ROWS
    return blocks[:count]


# 이전 버전과의 호환을 위해 유지 (기존에 실제로 만들어져 있던 12칸)
BLOCK_ROWS = generate_block_rows(KNOWN_GOOD_PAGES * 2)

PHOTO_EXTS = (".jpg", ".jpeg", ".png")

# 폴더명(세부설명)에 아래 키워드가 있으면 공종을 자동으로 추정합니다.
# 위에서부터 순서대로 검사하며, 먼저 걸리는 항목이 적용됩니다.
# 프로젝트 특성에 맞게 키워드나 공종명을 자유롭게 추가/수정하세요.
GONGJONG_KEYWORDS = [
    (["절삭"], "절삭공"),
    (["텍코팅", "택코팅", "프라임코팅", "코팅"], "택코팅공"),
    (["그레이팅", "측구", "배수"], "그레이팅설치공"),
    (["포설", "다짐", "아스콘", "포장"], "포장공"),
    (["구조물", "옹벽", "보도", "인도"], "구조물공"),
    (["표지", "안전시설"], "안전시설공"),
]


INVALID_FILENAME_CHARS = '\\/:*?"<>|'


def sanitize_filename(name, fallback="output"):
    """공사명 등 자유 텍스트를 윈도우 파일명으로 써도 안전하게 정리.
    파일명에 못 쓰는 문자(\\/:*?"<>|)는 지우고, 앞뒤 공백/마침표를 제거.
    정리 후 빈 문자열이면 fallback을 사용."""
    if not name:
        return fallback
    cleaned = "".join(ch for ch in name if ch not in INVALID_FILENAME_CHARS)
    cleaned = cleaned.strip(" .")
    return cleaned or fallback


def guess_gongjong(caption, default_gongjong=""):
    """폴더의 세부설명(caption) 텍스트에서 키워드를 찾아 공종을 추정합니다.
    일치하는 키워드가 없으면 default_gongjong을 그대로 돌려줍니다."""
    text = (caption or "").replace(" ", "")
    for keywords, gongjong in GONGJONG_KEYWORDS:
        for kw in keywords:
            if kw in text:
                return gongjong
    return default_gongjong


def list_stage_folders(photos_root, order=None):
    """photos_root 바로 아래의 하위 폴더를 전부 반환합니다 ('01_착공전' 처럼
    번호를 붙여도 되고, '착공전'처럼 번호 없이 이름만 써도 됩니다).
    '_resized'처럼 밑줄로 시작하는 폴더(내부용)는 제외합니다.

    order: [폴더이름, ...] 형태로 원하는 순서를 지정하면 그 순서대로 정렬하고,
           목록에 없는(새로 생긴) 폴더는 뒤에 이어붙입니다. 지정 안 하면
           이름(가나다/숫자) 순으로 정렬합니다."""
    if not photos_root or not os.path.isdir(photos_root):
        return []
    all_dirs = [
        f for f in sorted(glob.glob(os.path.join(photos_root, "*")))
        if os.path.isdir(f) and not os.path.basename(f).startswith("_")
    ]
    if order:
        by_base = {os.path.basename(f): f for f in all_dirs}
        ordered = [by_base[name] for name in order if name in by_base]
        remaining = [f for f in all_dirs if os.path.basename(f) not in order]
        return ordered + remaining
    return all_dirs


def list_photos(folder):
    """Windows는 대소문자를 구분하지 않아 '*.jpg'와 '*.JPG'가 같은 파일을 중복으로
    잡을 수 있으므로, 절대경로 기준으로 중복 제거 후 반환."""
    found = (
        glob.glob(os.path.join(folder, "*.jpg")) +
        glob.glob(os.path.join(folder, "*.jpeg")) +
        glob.glob(os.path.join(folder, "*.png")) +
        glob.glob(os.path.join(folder, "*.JPG")) +
        glob.glob(os.path.join(folder, "*.JPEG")) +
        glob.glob(os.path.join(folder, "*.PNG"))
    )
    seen = {}
    for f in found:
        seen[os.path.normcase(os.path.abspath(f))] = f
    return sorted(seen.values())


NUMERIC_PREFIX_RE = re.compile(r"^\d+_(.+)$")


def folder_caption(folder):
    """폴더명이 '01_착공전'처럼 숫자로 시작하면 숫자 부분만 떼어내고,
    번호 없이 '착공전'처럼 이름만 있으면 그대로 세부설명으로 씀."""
    base = os.path.basename(folder)
    m = NUMERIC_PREFIX_RE.match(base)
    label = m.group(1) if m else base
    return " ".join(list(label))


def dedupe_by_similarity(files):
    """perceptual hash로 서로 많이 닮은 사진을 걸러내고, 대표만 남긴 리스트를 반환."""
    if not HAS_IMAGEHASH or len(files) <= 1:
        return files
    kept, kept_hashes = [], []
    for f in files:
        try:
            h = imagehash.phash(PILImage.open(f))
        except Exception:
            kept.append(f)
            continue
        is_dup = any((h - kh) <= HASH_DUPLICATE_THRESHOLD for kh in kept_hashes)
        if not is_dup:
            kept.append(f)
            kept_hashes.append(h)
    return kept


def pick_representatives(files, n=2):
    """dedupe 후 남은 사진 중 대표 n장을 순서상 고르게 퍼진 위치에서 선택.
    n=1이면 가운데 사진, n=2면 처음/마지막(기존 방식), n=4 등 그 이상이면
    처음부터 끝까지 균등한 간격으로 n장을 뽑아 작업 진행 흐름이 골고루
    보이도록 함."""
    files = sorted(files)
    deduped = dedupe_by_similarity(files)
    if len(deduped) <= n:
        return deduped
    if n == 1:
        return [deduped[len(deduped) // 2]]
    if n == 2:
        return [deduped[0], deduped[-1]]
    last_idx = len(deduped) - 1
    indices = sorted(set(round(i * last_idx / (n - 1)) for i in range(n)))
    return [deduped[i] for i in indices]


# "사진대지" 페이지 제목이 반복되는 행 간격 (PAGE_ROWS와 동일, 2블록=1페이지마다 14행)
PAGE_HEADER_INTERVAL = PAGE_ROWS


def _resolve_anchor_cell(ws, cell):
    """cell이 병합된 범위의 일부(MergedCell)라면 그 범위의 좌상단 Cell을 찾아 반환
    (좌상단 셀만 실제로 값을 가질 수 있음). 병합돼 있지 않으면 그대로 반환.
    템플릿을 엑셀에서 직접 편집하다 보면 원래 비어있던 셀이 다른 범위와 병합되는
    경우가 있는데(예: 갑지 B39가 A39:H42 병합 범위에 들어가버림), 이런 템플릿도
    오류 없이 값을 채울 수 있도록 값을 쓰기 전에 항상 이 함수를 거치게 함."""
    if isinstance(cell, MergedCell):
        for mc in ws.merged_cells.ranges:
            if cell.coordinate in mc:
                return ws.cell(row=mc.min_row, column=mc.min_col)
    return cell


def _set_cell_value(ws, coord, value):
    """A1 표기 좌표(예: "A20")에 값을 씀."""
    _resolve_anchor_cell(ws, ws[coord]).value = value


def _set_cell(ws, row, col, value):
    """행/열 번호로 지정한 위치에 값을 씀."""
    _resolve_anchor_cell(ws, ws.cell(row=row, column=col)).value = value


def _merge_range_at(ws, row):
    """해당 행에서 시작하는 병합 범위를 찾아 반환 (없으면 None)."""
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row == row:
            return mc
    return None


def clone_page(ws, target_start_row):
    """1페이지(제목+사진 2장, 14행)의 서식·병합·행높이를 그대로 복제해서
    target_start_row 위치에 새로 만듭니다. 템플릿에 원래 있던 페이지 수(12장)
    보다 사진이 더 필요할 때, 이 함수로 페이지를 통째로 늘립니다.
    (값도 함께 복사되지만 실제 사용되는 블록은 build_album 본문에서 어차피
    새 값으로 덮어쓰므로 상관없음)"""
    src_start = PAGE_HEADER_FIRST_ROW
    offset = target_start_row - src_start
    if offset == 0:
        return

    # 1) 셀 값/서식 먼저 복사 (아직 병합 전이라 일반 Cell 상태일 때 처리)
    for row_offset in range(PAGE_ROWS):
        src_row = src_start + row_offset
        dst_row = target_start_row + row_offset

        src_dim = ws.row_dimensions.get(src_row)
        if src_dim is not None and src_dim.height is not None:
            ws.row_dimensions[dst_row].height = src_dim.height

        for col in range(1, 6):
            src_cell = ws.cell(row=src_row, column=col)
            dst_cell = ws.cell(row=dst_row, column=col)
            _resolve_anchor_cell(ws, dst_cell).value = src_cell.value
            dst_cell.font = copy(src_cell.font)
            dst_cell.border = copy(src_cell.border)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.number_format = src_cell.number_format

    # 2) 병합 범위는 마지막에 적용 (먼저 병합하면 위 반복문에서 MergedCell이 되어
    #    값을 못 씀)
    for mc in list(ws.merged_cells.ranges):
        if src_start <= mc.min_row < src_start + PAGE_ROWS:
            new_min_row = mc.min_row + offset
            new_max_row = mc.max_row + offset
            existing = _merge_range_at(ws, new_min_row)
            if existing is not None and existing.min_col == mc.min_col:
                ws.unmerge_cells(
                    start_row=existing.min_row, start_column=existing.min_col,
                    end_row=existing.max_row, end_column=existing.max_col,
                )
            ws.merge_cells(
                start_row=new_min_row, start_column=mc.min_col,
                end_row=new_max_row, end_column=mc.max_col,
            )


def ensure_capacity(ws, blocks_needed):
    """사진(블록)이 blocks_needed개 필요할 때, 페이지(14행)가 모자라면
    clone_page()로 필요한 만큼 페이지를 복제해서 뒤에 추가합니다."""
    pages_needed = -(-blocks_needed // 2)  # ceil(blocks_needed / 2)
    added_pages = 0
    for page_idx in range(KNOWN_GOOD_PAGES, pages_needed):
        target_start = PAGE_HEADER_FIRST_ROW + page_idx * PAGE_ROWS
        clone_page(ws, target_start)
        added_pages += 1
    return added_pages


def ensure_page_breaks(ws, last_row):
    """한 페이지(제목+사진 2장, 14행)가 다 차고도 여유 공간이 남아서, 엑셀이
    자동으로 다음 페이지의 제목 행을 현재 페이지 끝에 끌어다 붙이는 현상이
    있습니다 (제목+사진 2장이 한 세트로 안 나뉘고 다음 페이지 제목이 잘린 채
    걸쳐 보임). 각 "사진대지" 제목 행 바로 앞에 수동 페이지 나누기를 넣어
    페이지가 항상 제목 행에서 새로 시작하도록 고정합니다."""
    from openpyxl.worksheet.pagebreak import Break

    added = []
    row = PAGE_HEADER_FIRST_ROW + PAGE_HEADER_INTERVAL
    existing = {b.id for b in ws.row_breaks.brk}
    while row <= last_row:
        break_before = row - 1
        if break_before not in existing:
            ws.row_breaks.append(Break(id=break_before))
            existing.add(break_before)
            added.append(row)
        row += PAGE_HEADER_INTERVAL
    return added


def fix_missing_page_headers(ws, last_row):
    """일부 템플릿은 뒤쪽 예비 블록용 페이지 제목("사 진 대 지") 행이 비어있거나,
    심지어 병합 범위 자체가 다른 제목 행(A~E열 병합)과 다르게(B~E열만 병합되는
    식으로) 잘못된 경우가 있습니다 (사용된 적 없는 예비 칸이라 원본 제작자가
    제대로 만들어두지 않은 것). 맨 위 정상 제목 행을 기준으로, 병합 범위가
    다르면 바로잡고, 비어있으면 값/서식을 그대로 채워 넣습니다."""
    ref_row = PAGE_HEADER_FIRST_ROW
    ref_merge = _merge_range_at(ws, ref_row)
    ref_col = ref_merge.min_col if ref_merge else 1
    ref_cell = ws.cell(row=ref_row, column=ref_col)
    if not ref_cell.value:
        return []  # 기준이 될 정상 제목 셀 자체가 없으면 손대지 않음

    fixed_rows = []
    row = ref_row + PAGE_HEADER_INTERVAL
    while row <= last_row:
        cur_merge = _merge_range_at(ws, row)
        merge_mismatch = (
            ref_merge is not None and (
                cur_merge is None
                or cur_merge.min_col != ref_merge.min_col
                or cur_merge.max_col != ref_merge.max_col
            )
        )
        if merge_mismatch:
            if cur_merge is not None:
                ws.unmerge_cells(
                    start_row=cur_merge.min_row, start_column=cur_merge.min_col,
                    end_row=cur_merge.max_row, end_column=cur_merge.max_col,
                )
            ws.merge_cells(
                start_row=row, start_column=ref_merge.min_col,
                end_row=row, end_column=ref_merge.max_col,
            )

        cell = ws.cell(row=row, column=ref_col)
        needs_value_fix = not cell.value

        # 병합 범위 안의 모든 열(A~E 등)에 테두리/배경까지 기준 행과 동일하게 맞춤
        # (병합된 셀이라도 테두리는 열마다 따로 저장되므로 전체 열을 다 맞춰야 함)
        min_col = ref_merge.min_col if ref_merge else ref_col
        max_col = ref_merge.max_col if ref_merge else ref_col
        border_fixed = False
        for col in range(min_col, max_col + 1):
            ref_c = ws.cell(row=ref_row, column=col)
            cur_c = ws.cell(row=row, column=col)
            if (cur_c.border.top.style, cur_c.border.bottom.style,
                    cur_c.border.left.style, cur_c.border.right.style) != (
                    ref_c.border.top.style, ref_c.border.bottom.style,
                    ref_c.border.left.style, ref_c.border.right.style) or \
                    (cur_c.fill.fgColor.rgb if cur_c.fill.fgColor else None) != \
                    (ref_c.fill.fgColor.rgb if ref_c.fill.fgColor else None):
                cur_c.border = copy(ref_c.border)
                cur_c.fill = copy(ref_c.fill)
                border_fixed = True

        if needs_value_fix:
            _resolve_anchor_cell(ws, cell).value = ref_cell.value
            cell.font = copy(ref_cell.font)
            cell.alignment = copy(ref_cell.alignment)

        if needs_value_fix or merge_mismatch or border_fixed:
            fixed_rows.append(row)

        row += PAGE_HEADER_INTERVAL
    return fixed_rows


LABEL_COLUMNS = (1, 2, 3, 4, 5)  # A~E열


def fix_block_label_styles(ws, used_blocks):
    """일부 템플릿은 뒤쪽 예비 블록(사용된 적 없는 칸)의 '공사명/공종/세부설명'
    라벨(A열 텍스트)이 비어있거나, 글꼴·정렬·테두리가 앞쪽 블록과 다른 경우가
    있습니다. 첫 번째 블록(항상 정상적으로 채워져 있음)을 기준으로 삼아,
    다른 블록에서 기준과 다르게 생긴 셀을 찾아 라벨 텍스트와 서식을 그대로
    복사해 맞춥니다."""
    if len(used_blocks) < 2:
        return []

    ref_block = used_blocks[0]
    fixed_rows = []

    for block in used_blocks[1:]:
        for role in range(3):  # 0=공사명행, 1=공종행, 2=세부설명행
            ref_row = ref_block[role]
            row = block[role]
            if row == ref_row:
                continue

            ref_label_cell = ws.cell(row=ref_row, column=1)
            cur_label_cell = ws.cell(row=row, column=1)
            row_fixed = False

            if not cur_label_cell.value and ref_label_cell.value:
                _resolve_anchor_cell(ws, cur_label_cell).value = ref_label_cell.value
                row_fixed = True

            for col in LABEL_COLUMNS:
                ref_cell = ws.cell(row=ref_row, column=col)
                cur_cell = ws.cell(row=row, column=col)
                ref_border = (ref_cell.border.top.style, ref_cell.border.bottom.style,
                              ref_cell.border.left.style, ref_cell.border.right.style)
                cur_border = (cur_cell.border.top.style, cur_cell.border.bottom.style,
                              cur_cell.border.left.style, cur_cell.border.right.style)
                ref_fill = ref_cell.fill.fgColor.rgb if ref_cell.fill.fgColor else None
                cur_fill = cur_cell.fill.fgColor.rgb if cur_cell.fill.fgColor else None
                style_mismatch = (
                    cur_cell.font.name != ref_cell.font.name
                    or cur_cell.font.b != ref_cell.font.b
                    or cur_cell.font.sz != ref_cell.font.sz
                    or cur_cell.alignment.horizontal != ref_cell.alignment.horizontal
                    or cur_cell.alignment.vertical != ref_cell.alignment.vertical
                    or cur_border != ref_border
                    or cur_fill != ref_fill
                )
                if style_mismatch:
                    cur_cell.font = copy(ref_cell.font)
                    cur_cell.alignment = copy(ref_cell.alignment)
                    cur_cell.border = copy(ref_cell.border)
                    cur_cell.fill = copy(ref_cell.fill)
                    cur_cell.number_format = ref_cell.number_format
                    row_fixed = True

            if row_fixed:
                fixed_rows.append(row)

    return fixed_rows


# 실제로 인쇄까지 확인된 '대평2리 마을안길 포장공사.xlsx' 사진대지 파일의 인쇄 설정
# 기준값. 템플릿마다 인쇄 배율/여백이 제각각이라 그대로 출력하면 사진이 페이지
# 경계에 걸리거나 여백이 안 맞을 수 있어, 매 생성마다 이 값으로 맞춰줍니다.
PRINT_SETUP = {
    "갑지": {
        "fit_to_page": True,
        "margins": dict(left=0.94, right=0.76, top=1.0, bottom=1.0, header=0.51, footer=0.51),
    },
    "공사사진": {
        "scale": 110,
        "margins": dict(left=0.75, right=0.43, top=0.84, bottom=0.43, header=0.28, footer=0.24),
    },
}


def apply_print_page_setup(wb):
    """갑지/공사사진 시트에 PRINT_SETUP 기준값(용지 A4·세로, 배율/여백)을 적용."""
    for name, setup in PRINT_SETUP.items():
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        ws.page_setup.paperSize = 9  # A4
        ws.page_setup.orientation = "portrait"
        if setup.get("fit_to_page"):
            ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        if "scale" in setup:
            ws.page_setup.scale = setup["scale"]
        m = setup["margins"]
        ws.page_margins.left = m["left"]
        ws.page_margins.right = m["right"]
        ws.page_margins.top = m["top"]
        ws.page_margins.bottom = m["bottom"]
        ws.page_margins.header = m["header"]
        ws.page_margins.footer = m["footer"]


def build_stage_list(photos_root, folder_order=None):
    """물리 폴더를 그대로 1폴더=1단계로 사용하는 기본 stage_list를 만듭니다.
    각 stage: {"key": 폴더명, "folder": 폴더경로, "caption": 세부설명}"""
    folders = list_stage_folders(photos_root, order=folder_order)
    return [
        {"key": os.path.basename(f), "folder": f, "caption": folder_caption(f)}
        for f in folders
    ]


def build_album(template_path, output_path, project_name, gongjong, owner,
                 photos_root, manual_picks=None, gongjong_by_folder=None,
                 per_folder=2, photo_count_by_folder=None, folder_order=None,
                 stage_list=None, resized_dir=None, log=print):
    """
    stage_list: [{"key": ..., "folder": ..., "caption": ...}, ...] 처럼 "단계"를
                직접 지정하고 싶을 때 사용. 폴더 하나를 여러 세부 단계로 나눠서
                쓰고 싶을 때(예: '포장공' 폴더 안 사진들을 '노면절삭'/'택코팅'
                등으로 나눔) 유용함 - 이때 여러 stage가 같은 folder를 공유할 수
                있고, key만 서로 다르면 됨(예: "포장공::노면절삭").
                지정 안 하면 photos_root 아래 폴더를 그대로 1폴더=1단계로 사용.
    manual_picks: {"01_착공전": ["파일1.jpg", "파일2.jpg"], ...}
                  stage_list를 안 쓰면 폴더명, 쓰면 stage의 key를 키로 사용.
                  값이 파일명이면 folder 안에서 찾고, 절대경로면 그대로 사용.
                  (여기 지정된 개수가 곧 그 단계의 사진 장수가 됨. 2장/4장 모두 가능)
    gongjong_by_folder: {"01_착공전": "절삭공", ...} (키는 manual_picks와 동일한 규칙)
                  단계별로 공종을 직접 지정하고 싶을 때 사용. 지정 안 한 단계는
                  세부설명에서 키워드를 찾아 자동 추정하고, 그마저 안 되면 gongjong
                  (기본 공종) 값을 사용합니다.
    photo_count_by_folder: {"01_착공전": 4, ...} (키는 manual_picks와 동일한 규칙)
                  자동 선택 시 단계별로 몇 장을 뽑을지 지정 (2 또는 4). 지정 안 한
                  단계는 per_folder(기본 2장)를 사용. manual_picks가 있으면
                  이 값과 무관하게 manual_picks 개수를 그대로 사용.
    folder_order: ["착공전", "노면절삭", ...] 처럼 폴더 처리 순서를 직접
                  지정하고 싶을 때 사용 (stage_list를 안 쓸 때만 적용됨).
    반환값: [(사진경로, 세부설명caption, 공종), ...] 실제로 사용된 사진 목록
    """
    manual_picks = manual_picks or {}
    gongjong_by_folder = gongjong_by_folder or {}
    photo_count_by_folder = photo_count_by_folder or {}
    resized_dir = resized_dir or os.path.join(photos_root, "_resized")

    if stage_list is None:
        stage_list = build_stage_list(photos_root, folder_order=folder_order)
    if not stage_list:
        raise RuntimeError(
            f"'{photos_root}' 폴더 아래에서 단계별 사진 폴더를 찾지 못했습니다."
        )

    photo_items = []
    for stage in stage_list:
        key = stage["key"]
        folder = stage["folder"]
        caption = stage["caption"]
        all_files = list_photos(folder)

        manual_names = manual_picks.get(key)
        chosen = None
        is_manual = False
        if manual_names:
            candidates = []
            for name in manual_names:
                p = name if os.path.isabs(name) else os.path.join(folder, name)
                if os.path.isfile(p):
                    candidates.append(p)
            if candidates:
                chosen = candidates
                is_manual = True

        if not chosen:
            count = photo_count_by_folder.get(key, per_folder)
            chosen = pick_representatives(all_files, n=count)

        stage_gongjong = gongjong_by_folder.get(key)
        gj_tag = "지정"
        if not stage_gongjong:
            stage_gongjong = guess_gongjong(caption, gongjong)
            gj_tag = "자동추정"

        tag = "수동 지정" if is_manual else "자동 선택"
        log(f"[{key}] 사진 {tag} {len(chosen)}장, 공종 {gj_tag}: '{stage_gongjong}' "
            f"| {[os.path.basename(f) for f in chosen]}")
        for f in chosen:
            photo_items.append((f, caption, stage_gongjong))

    if not photo_items:
        raise RuntimeError("선택된 사진이 한 장도 없습니다.")

    MAX_PHOTOS_SANITY = 300  # 실수로 지나치게 큰 값이 들어오는 것을 막는 안전 한도
    if len(photo_items) > MAX_PHOTOS_SANITY:
        raise RuntimeError(
            f"사진이 {len(photo_items)}장으로 너무 많습니다 (최대 {MAX_PHOTOS_SANITY}장).\n"
            f"폴더/장수 설정을 확인해주세요."
        )

    wb = openpyxl.load_workbook(template_path)
    if "갑지" not in wb.sheetnames or "공사사진" not in wb.sheetnames:
        raise RuntimeError("템플릿 파일에 '갑지', '공사사진' 시트가 없습니다. 올바른 템플릿인지 확인하세요.")

    ws1 = wb["갑지"]
    _set_cell_value(ws1, "A20", project_name)
    _set_cell_value(ws1, "B39", owner)

    ws2 = wb["공사사진"]
    ws2._images = []

    added_pages = ensure_capacity(ws2, len(photo_items))
    if added_pages:
        log(f"템플릿 기본 용량(12장)을 넘어서 페이지 {added_pages}개를 추가로 복제했습니다.")

    used_blocks = generate_block_rows(len(photo_items))
    last_used_last_row = used_blocks[-1][2]
    if ws2.max_row > last_used_last_row:
        ws2.delete_rows(last_used_last_row + 1, ws2.max_row - last_used_last_row)

    os.makedirs(resized_dir, exist_ok=True)

    for i, ((cname, gname, dname, prow), (src_path, caption, item_gongjong)) in enumerate(zip(used_blocks, photo_items)):
        _set_cell(ws2, cname, 3, project_name)
        _set_cell(ws2, gname, 3, item_gongjong)
        _set_cell(ws2, dname, 3, caption)
        _set_cell(ws2, gname, 1, " 공         종 ")

        im = ImageOps.exif_transpose(PILImage.open(src_path)).convert("RGB")
        im2 = im.resize((IMG_W_PX, IMG_H_PX), PILImage.LANCZOS)
        out_path = os.path.join(resized_dir, f"b{i+1:02d}.jpg")
        im2.save(out_path, quality=88)

        xl_img = XLImage(out_path)
        xl_img.width, xl_img.height = IMG_W_PX, IMG_H_PX
        r0 = prow - 1
        frm = AnchorMarker(col=1, colOff=0, row=r0, rowOff=0)
        to = AnchorMarker(col=3, colOff=BOX_WIDTH_EMU, row=r0, rowOff=BOX_HEIGHT_EMU)
        xl_img.anchor = TwoCellAnchor(_from=frm, to=to, editAs="oneCell")
        ws2.add_image(xl_img)

    last_row = used_blocks[-1][2]
    ws2.print_area = f"A1:E{last_row}"

    label_fixed = fix_block_label_styles(ws2, used_blocks)
    if label_fixed:
        log(f"공사명/공종/세부설명 라벨 서식이 깨져있던 {len(label_fixed)}행을 복구했습니다: {label_fixed}")

    header_fixed = fix_missing_page_headers(ws2, last_row)
    if header_fixed:
        log(f"페이지 제목이 비어있던 {len(header_fixed)}행을 채웠습니다: {header_fixed}")

    breaks_added = ensure_page_breaks(ws2, last_row)
    if breaks_added:
        log(f"페이지 나누기를 제목 행 앞에 고정했습니다: {breaks_added}")

    apply_print_page_setup(wb)
    log("인쇄 페이지 크기/배율/여백을 기준 서식(대평2리 사진대지 기준)으로 맞췄습니다.")

    wb.save(output_path)
    return photo_items
